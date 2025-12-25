import os
import json
import pandas as pd
from common.logger import get_logger
from pathlib import Path
from typing import Dict, Any 

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

logger = get_logger("agent.job.saver")

def job_to_dict(job):
    """Chuyển NormalizedJob -> dict để lưu JSON/Excel"""
    return {
        "title": job.title,
        "canonical_title": job.canonical_title,
        "company": job.company,
        "description": job.description,
        "cleaned_text": job.cleaned_text,
        "skills": job.skills,
        "experience": job.experience,
        "location": job.location,
        "platform": job.platform,
        "source_url": job.raw_ref,
        "debug": job.debug,
    }
    
def center_column_by_header(ws, header_name):
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == header_name:
            for row in ws.iter_rows(min_col=idx, max_col=idx, min_row=2):
                row[0].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
            break

    
# Lưu dữ liệu đã chuẩn hóa ra file Excel
def save_to_excel(jobs, excel_path: Path | None = None) -> Path:
    
    # Đường dẫn lưu job
    if excel_path is None:
        excel_path = Path("data") / "jobs" / "normalized_jobs.xlsx"

    excel_path.parent.mkdir(parents=True, exist_ok=True)

    job_dicts = [job_to_dict(j) for j in jobs]
    
    # Job data cần lưu để xem
    rows = []
    for j in job_dicts:
        locs = j.get("location", [])

        location_raw = "\n".join(
            loc.get("raw") for loc in locs if isinstance(loc, dict) and loc.get("raw")
        )
        
        ct = j.get("canonical_title", {})
        rows.append({
            "Title": j.get("title"),
            "Role": " ".join(
                            part for part in [ct.get("seniority"), ct.get("role")] if part),
            "Company": j.get("company"),
            "Platform": j.get("platform"),
            "Location": location_raw
        })

    df = pd.DataFrame(rows)

    df.to_excel(excel_path, sheet_name="jobs", index=False)
    
    wb = load_workbook(excel_path)
    ws = wb["jobs"]

    # 1. In đậm header + căn giữa
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    # 2. Wrap text + căn trái giữa cho toàn bộ
    content_alignment = Alignment(
        wrap_text=True,
        vertical="center",
        horizontal="left"
    )
    
    # căn giữa cho cột Platfrom
    center_column_by_header(ws, "Platform")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = content_alignment

    # 3. Set độ rộng cột (rất quan trọng)
    column_widths = {
        "A": 30,  # Title
        "B": 20,  # Role
        "C": 25,  # Company
        "D": 15,  # Platform
        "E": 45,  # Location (nhiều text)
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 4. Bọc toàn bộ bằng Table (Excel Table đúng nghĩa)
    end_row = ws.max_row
    end_col = ws.max_column

    table = Table(
        displayName="JobsTable",
        ref=f"A1:{chr(64 + end_col)}{end_row}"
    )

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(excel_path)

    return excel_path

# Lưu dữ liệu đã chuẩn hóa ra file Json
def to_serializable(obj):
    if isinstance(obj, list):
        return [to_serializable(o) for o in obj]
    elif isinstance(obj, dict):
        return {k: to_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, datetime):
        return obj.isoformat()
    elif hasattr(obj, "to_dict"):
        return to_serializable(obj.to_dict())
    elif hasattr(obj, "__dict__"):
        return to_serializable(vars(obj))
    else:
        return obj

def save_to_json(jobs, json_path: Path | None = None) -> Path:
    if json_path is None:
        json_path = Path("data") / "jobs" / "normalized_jobs.json"
    json_path.parent.mkdir(parents=True, exist_ok=True)

    # Convert tất cả object phức tạp sang dict/string
    jobs_serializable = to_serializable(jobs)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(jobs_serializable, f, ensure_ascii=False, indent=2)

    return json_path

def save_jobs(jobs):
    logger.info("Start Save")

    json_path = save_to_json(jobs)
    # excel_path = save_to_excel(jobs)

    logger.info(f"Extraction completed")
    logger.info(f"Saved JSON  : {json_path}")
    # logger.info(f"Saved Excel : {excel_path}")

    return jobs

# def save_jobs(jobs,
#               json_path=os.path.join("data", "jobs", "normalized_jobs.json"),
#               excel_path=os.path.join("data", "jobs", "normalized_jobs.xlsx")):

#     os.makedirs(os.path.dirname(json_path), exist_ok=True)
#     os.makedirs(os.path.dirname(excel_path), exist_ok=True)

#     dicts = [job_to_dict(j) for j in jobs]

#     with open(json_path, "w", encoding="utf-8") as f:
#         json.dump(dicts, f, ensure_ascii=False, indent=2)

#     df = pd.DataFrame(dicts)
#     df.to_excel(excel_path, index=False)

#     logger.info(f"Saved {len(jobs)} jobs to {json_path} and {excel_path}")

